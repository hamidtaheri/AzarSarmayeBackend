import datetime
from random import randrange

import jdatetime
import openpyxl
from crum import get_current_user
from django.contrib.auth import get_user_model
from django.contrib.auth.models import AbstractUser
from django.contrib.contenttypes.fields import GenericForeignKey, GenericRelation
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import PermissionDenied
from django.core.validators import MaxValueValidator, MinValueValidator
from django.db import models
from django.db.models import QuerySet, Sum, Q
from django.utils.timezone import now


class MyException(Exception):
    pass


day_in_month = 30
day_for_calculate_presenter_profit = 180  # تعدادروزی که پس از آن سود معرفی به معرف تعلق نمیگیرد


def get_storage_path(instance, filename):
    return f"{instance.content_type.model}/{instance.object_id}_{randrange(100, 999)}_{filename}"


class User(AbstractUser):
    mobile = models.CharField('تلفن همراه', max_length=11, blank=True, null=True, help_text='شماره موبایل')

    def __str__(self):
        return f'{self.username}'

    permissions = (
        ("can_add_user", "میتواند کاربر جدید ایجاد کند"),
    )


class ImageKind(models.Model):
    name = models.CharField(max_length=100, default='')
    description = models.TextField(null=True, blank=True)

    def __str__(self):
        return self.name


class Image(models.Model):
    """
    کلاسی که برای همه تصاویر در سایر کلاس ها استفاده میشود.
    """
    description = models.CharField(max_length=500, null=True, blank=True)
    image = models.ImageField(upload_to=get_storage_path)
    kind = models.ForeignKey(ImageKind, on_delete=models.SET_NULL, null=True, blank=True)

    content_type = models.ForeignKey(ContentType, on_delete=models.CASCADE)
    object_id = models.PositiveIntegerField()
    content_object = GenericForeignKey()

    # logging fields
    created = models.DateTimeField(auto_now_add=True, editable=False)
    created_by = models.ForeignKey(get_user_model(), on_delete=models.DO_NOTHING, blank=True, null=True,
                                   related_name='image_create_by', editable=False)
    modified = models.DateTimeField(auto_now=True, editable=False)
    modified_by = models.ForeignKey(get_user_model(), on_delete=models.DO_NOTHING, blank=True, null=True,
                                    related_name='image_modified_by', editable=False)

    class Meta:
        verbose_name = 'تصویر'
        verbose_name_plural = 'تصویر ها'

    def __str__(self):
        image_type = ContentType.objects.get(app_label='api', model=self.content_type.model)
        image_type.get_object_for_this_type(id=self.object_id)
        return f"{image_type.get_object_for_this_type(id=self.object_id)}"


class TransactionKind(models.Model):
    title = models.CharField(verbose_name='نوع تراکنش', max_length=30)
    description = models.TextField(verbose_name='توضیح')

    def __str__(self):
        return f'{self.id}-{self.title}'


class Transaction_old(models.Model):
    charge = 'charge'
    withdraw = 'withdraw'
    withdraw_profit = 'withdraw_profit'

    TRASACTION_KINDS = [
        (charge, "واریز"),
        (withdraw, "برداشت"),
        (withdraw_profit, "برداشت سود"),
    ]

    user = models.ForeignKey(get_user_model(), verbose_name='کاربر', on_delete=models.DO_NOTHING,
                             related_name='transactions', blank=False, null=False)
    # date = jmodels.jDateField(verbose_name='تاریخ', blank=False, null=False)
    date = models.DateField(verbose_name='تاریخ', blank=False, null=False)
    amount = models.PositiveBigIntegerField(verbose_name='مبلغ', blank=False, null=False)
    kind = models.ForeignKey(TransactionKind, verbose_name='نوع', on_delete=models.CASCADE)

    @property
    def shamse_date(self):
        return jdatetime.date.fromgregorian(date=self.date).strftime(format="%Y-%m-%d")

    class Meta:
        verbose_name = "عملیات مالی(واریز/برداشت)"
        verbose_name_plural = "عملیات مالی(واریز/برداشت)"

        permissions = (
            ("can_add_transaction_for_self", "میتواند برای خودش تراکنش ثبت کند"),
            ("can_add_transaction_for_all", "میتواند برای دیگران تراکنش ثبت کند"),
            ("can_view_transaction_for_all", "میتواند تراکنش دیگران را مشاهده کند"),
        )

    def __str__(self):
        return f'{self.kind} {self.amount} {self.user} {self.date}'

    def before_save_transaction(self):
        current_user: User = get_current_user()
        # کاربر وارد نشده
        if (not current_user.is_authenticated) or current_user.is_anonymous or current_user is None:
            raise PermissionDenied()
        if not current_user.has_perm('can_add_transaction_for_all'):  # کابر جاری مجاز نیست برای دیگران تراکنش ثبت کند
            if self.user != current_user:
                raise PermissionDenied()
        if not self.pk:
            self.created_by = current_user
        self.modified_by = current_user

        print(self)

    def save(self, *args, **kwargs):
        print(f'save {self.user} {self.kind} {self.amount}')

        self.before_save_transaction()

        profit_calculate_objects_for_user: QuerySet[ProfitCalculate] = ProfitCalculate.objects.filter(user=self.user)
        # قبلا در جدول محاسبه سود برای این کاربر رکوردی ثبت نشده است
        new_profit_calculate: ProfitCalculate = ProfitCalculate()
        if profit_calculate_objects_for_user.count() == 0 or profit_calculate_objects_for_user is None:
            new_profit_calculate.user = self.user
            new_profit_calculate.date_from = self.date
            new_profit_calculate.amount = self.amount

            super().save(*args, **kwargs)

            new_profit_calculate.transaction = self
            new_profit_calculate.save()
            # اولین رکورد در جدول محاسبه سود برای این کاربر ثبت شد

        else:
            # به روز رسانی آخرین رکورد جدول محاسبه سود و ایجاد رکورد جدید
            profit_calculate_for_user_latest_by_created = profit_calculate_objects_for_user.latest('created')
            profit_calculate_for_user_latest_by_id = profit_calculate_objects_for_user.latest('id')
            profit_calculate_for_user_latest_date_to = profit_calculate_objects_for_user.filter(
                date_to__isnull=True)
            if profit_calculate_for_user_latest_date_to.count() == 1 and \
                    profit_calculate_for_user_latest_by_created == profit_calculate_for_user_latest_by_id and \
                    profit_calculate_for_user_latest_by_id == profit_calculate_for_user_latest_date_to[0]:
                latest_profit_calculate: ProfitCalculate = profit_calculate_for_user_latest_by_created
                latest_profit_calculate.date_to = self.date
                days = (self.date - latest_profit_calculate.date_from).days
                if days < 0:
                    raise MyException(
                        f'new transaction date ({self.date}) must newest from last transaction ({latest_profit_calculate.date_from})')
                latest_profit_calculate.days = days
                latest_profit_calculate.save()

                new_profit_calculate.user = self.user
                new_profit_calculate.date_from = self.date
                new_amount = 0
                if self.type == 'withdraw':  # برداشت از حساب
                    new_amount = latest_profit_calculate.amount - self.amount
                elif self.type == 'charge':  # واریز به حساب
                    new_amount = latest_profit_calculate.amount + self.amount

                new_profit_calculate.amount = new_amount

                super().save(*args, **kwargs)

                new_profit_calculate.transaction = self
                new_profit_calculate.save()

            print(f'ProfitCalculate  availebale for user:{self.user}')


class Profile(models.Model):
    user = models.OneToOneField(User, on_delete=models.CASCADE, related_name='profile', blank=True, null=True, )
    first_name = models.CharField(max_length=100, blank=True, null=True)
    last_name = models.CharField(max_length=100, blank=True, null=True)
    code_meli = models.CharField(max_length=10, blank=True, null=True)
    adress = models.CharField(max_length=500, blank=True, null=True)
    shomare_kart = models.CharField(max_length=100, blank=True, null=True)
    shomare_hesab = models.CharField(max_length=100, blank=True, null=True)
    presenter = models.ForeignKey('self', verbose_name='معرف', blank=True, null=True,
                                  on_delete=models.CASCADE, related_name='moarefi_shode_ha')  # moarefi_shode_ha
    percent = models.IntegerField(blank=True, null=True)
    presenter_percent = models.IntegerField(blank=True, null=True)
    get_profit = models.BooleanField(blank=True, null=True)
    description = models.CharField(max_length=100, blank=True, null=True)
    charge_to_presenter = models.BooleanField(blank=True, null=True)
    self_presenter = models.BooleanField(blank=True, null=True)
    tel = models.CharField(max_length=20, blank=True, null=True)
    mobile1 = models.CharField(max_length=11, blank=True, null=True)
    mobile2 = models.CharField(max_length=11, blank=True, null=True)
    presenter_2 = models.IntegerField(blank=True, null=True)
    presenter_percent_2 = models.IntegerField(blank=True, null=True)
    self_presenter_2 = models.BooleanField(blank=True, null=True)
    images = GenericRelation(Image, related_name='profile_images')

    class Meta:
        verbose_name = "پروفایل"
        verbose_name_plural = "پروفایل"

        permissions = (
            ("can_add_profile", "ایجاد پروفایل"),
            ("view_all_profiles", " مشاهده همه پروفایل ها"),
            ("can_edit_profile_for_self", "ویرایش پروفایل خودش"),
            ("can_edit_profile_for_all", "ویرایش پروفایل دیگران"),
            ("can_delete_profile_for_all", "حذف پروفایل"),
        )

    @property
    def first_transaction(self):
        """اولین سپرده گزاری"""
        tr = Transaction.objects.filter(profile=self, kind_id=1).order_by("effective_date").first()
        return tr

    def tarakonesh_sum_ta(self, kind: int, ta: datetime = None, ) -> float:
        if ta:
            tr = Transaction.objects.filter(profile=self, kind=kind, effective_date__lte=ta).aggregate(
                Sum('amount'))
        else:
            tr = Transaction.objects.filter(profile=self, kind=kind).aggregate(Sum('amount'))

        r = tr['amount__sum'] or 0
        return r

    def mojodi_ta(self, ta: datetime.date = datetime.date.today()) -> float:
        """
        موجودی حساب یعنی مجموع سرمایه گزاری ها و سودها منهای برداشت سود و مرجوعی
        :param ta:  تا تاریخ اگر تاریخ
        :return:  موجودی
        """
        if ta is None:
            ta = datetime.date.today()

        seporde = self.tarakonesh_sum_ta(kind=1, ta=ta)  # سپرده گذاری
        marjo = self.tarakonesh_sum_ta(kind=2, ta=ta)  # مرجوع
        variz_sod = self.tarakonesh_sum_ta(kind=3, ta=ta)  # واریز سود
        bardasht_sod = self.tarakonesh_sum_ta(kind=4, ta=ta)  # برداشت سود
        variz_sod_moarefi = self.tarakonesh_sum_ta(kind=5, ta=ta)  # واریز سود معرفی
        r = (seporde + variz_sod + variz_sod_moarefi) - (marjo + bardasht_sod)
        return r

    def mojodi_moarefishodeha_ta(self, ta_date: datetime.date = datetime.date.today()) -> float:
        """
        محاسبه جمع موجودی معرفی شده ها در بازه از n روز پیش تاکنون
        n= day_for_calculate_presenter_profit
        """
        transactions: QuerySet[Transaction] = Transaction.objects.filter(profile__presenter=self,
                                                                         effective_date__lte=ta_date,
                                                                         effective_date__gte=
                                                                         ta_date - datetime.timedelta(
                                                                             day_for_calculate_presenter_profit)
                                                                         )
        seporde = transactions.filter(kind=1).aggregate(seporde=Sum('amount'))['seporde'] or 0
        marjo = transactions.filter(kind=2).aggregate(marjo=Sum('amount'))['marjo'] or 0
        variz_sod = transactions.filter(kind=3).aggregate(variz_sod=Sum('amount'))['variz_sod'] or 0
        bardasht_sod = transactions.filter(kind=4).aggregate(bardasht_sod=Sum('amount'))['bardasht_sod'] or 0
        variz_sod_moarefi = transactions.filter(kind=5).aggregate(variz_sod_moarefi=Sum('amount'))[
                                'variz_sod_moarefi'] or 0
        r = (seporde + variz_sod + variz_sod_moarefi) - (marjo + bardasht_sod)
        return r

    def mohasebe_sod_old(self, az_date: datetime.date, ta_date: datetime.date):
        """
        محاسبه سود بر اساس اینکه هر واریزی درصد سود خود را دارد
        :param user:
        :param az_date:
        :param ta_date:
        :return:لیست محاسبات سود و مجموع انها را بر میگرداند
        """
        sum_of_sod = 0
        # واریزی های کاربر تا پیش از تاریخ پایان

        mohasebat_sod: list[ProfitCalculate] = list[ProfitCalculate]()
        tarakoneshs: QuerySet[Transaction] = Transaction.objects.filter(Q(kind_id=1) | Q(kind_id=2),
                                                                        profile=self,
                                                                        effective_date__lte=ta_date, )
        for tr in tarakoneshs:
            mohasebe_sod: ProfitCalculate = tr.profit_calculator(az_date=az_date, ta_date=ta_date)
            sum_of_sod = sum_of_sod + mohasebe_sod.calculated_amount
            mohasebat_sod.append(mohasebe_sod)

        return mohasebat_sod, sum_of_sod

    def mohasebe_sod_moarefi(self, az_date: datetime, ta_date: datetime):
        sum_of_sod = 0
        mohasebat_sod: list[ProfitCalculate] = list[ProfitCalculate]()
        # پیدا کردن تمام سپرده گزاری هایی که این پروفایل(self)معرف آنهاست و تاریخ موثر آنها کوچکتر مساوی تا است
        tarakoneshs: QuerySet[Transaction] = Transaction.objects.filter(profile__presenter=self,
                                                                        effective_date__lte=ta_date,
                                                                        kind_id=1)
        for tr in tarakoneshs:
            mohasebe_sod: ProfitCalculate = tr.presenter_profit_calculator(az_date=az_date, ta_date=ta_date)
            sum_of_sod = sum_of_sod + mohasebe_sod.calculated_amount
            mohasebat_sod.append(mohasebe_sod)

        return mohasebat_sod, sum_of_sod

    def __str__(self):
        return f'{self.first_name}   {self.last_name}'


class ProfileImageGallery(models.Model):
    image_kind = models.ForeignKey(ImageKind, on_delete=models.SET_NULL, null=True, blank=True)
    image = GenericRelation(Image, related_name='profile_images')
    profile = models.ForeignKey(Profile, on_delete=models.SET_NULL, related_name='profile_images', null=True,
                                blank=True)


class Transaction(models.Model):
    profile = models.ForeignKey(Profile, on_delete=models.CASCADE, related_name='transactions')
    tarikh = models.CharField(max_length=10)
    date = models.DateField(null=True, blank=True, )
    Tarikh_Moaser = models.CharField(max_length=10)
    effective_date = models.DateField(null=True, blank=True)
    expire_date = models.DateField(null=True, blank=True, verbose_name='تاریخ انتهای قرارداد')
    Tarikh_Moaser_Moaref = models.CharField(max_length=10)
    presenter_effective_date = models.DateField(null=True, blank=True)
    date_time = models.DateTimeField()
    amount = models.BigIntegerField()
    kind = models.ForeignKey(TransactionKind, on_delete=models.CASCADE, related_name='transactions')
    NahveyePardakht = models.CharField(max_length=40, blank=True, null=True)
    description = models.TextField(blank=True, null=True)
    Tbl_Pardakht_List_id = models.IntegerField(blank=True, null=True)
    percent = models.IntegerField()
    # صرفا برای محاسبه مهر۱۴۰۰ برای اینکه محاسبه سود معرف به شیوه قدیمی محاسبه شود
    DarMelyoon_Moaref = models.IntegerField(blank=True, null=True)
    images = GenericRelation(Image, related_name='transaction_images')


    # logging fields
    created = models.DateTimeField(auto_now_add=True, editable=False)
    created_by = models.ForeignKey(get_user_model(), on_delete=models.DO_NOTHING, blank=False, null=False,
                                   related_name='transactions_create_by', editable=False)
    modified = models.DateTimeField(auto_now=True, editable=False)
    modified_by = models.ForeignKey(get_user_model(), on_delete=models.DO_NOTHING, blank=True, null=True,
                                    related_name='transactions_modified_by', editable=False)

    def __str__(self):
        return f'{self.id} {self.profile.first_name} {self.profile.last_name}({self.profile.id})'

    class Meta:
        ordering = ['effective_date']
        permissions = (
            ("view_all_transactions", "مشاهده همه تراکنش ها "),
            ("add_transaction_for_all", "میتواند برای دیگران تراکنش ثبت کند"),

        )

    def percent_calculator(self) -> float:
        """
        محاسبه درصد سود سرمایه گزاری بر اساس پلکان
        """
        if self.effective_date <= sh2m('1400/04/01'):
            return self.percent
        else:
            mojodi = self.profile.mojodi_ta(ta=self.effective_date)
            #   از کوچکتر مساوی از موجودی و تا بزرگتر از موجودی
            # lte=Less Than or Equal   gt=Greater Than
            dar_melyon = Pelekan.objects.get(kind_id=1, az__lte=mojodi, ta__gt=mojodi).percent
            return dar_melyon

    def presenter_percent_calculator(self, ta_date: datetime.date) -> float:
        """
        محاسبه درصد سود معرف بر اساس پلکان
        """
        if self.effective_date <= sh2m('1400/07/30'):
            # محاسبه سود بر اساس درصد ثبت شده در رکورد تراکنش
            return self.DarMelyoon_Moaref
        else:
            mojodi = self.profile.mojodi_ta(ta=self.effective_date)
            mojodi_moarefishodeha = self.profile.presenter.mojodi_moarefishodeha_ta(ta_date=ta_date)
            mojodi_kol = mojodi + mojodi_moarefishodeha
            # محاسبه درصد سود معرف بر اساس پلکان
            percent = Pelekan.objects.get(kind_id=2, az__lt=mojodi_kol, ta__gte=mojodi_kol).percent
            return percent

    def profit_calculator(self, az_date: datetime.date, ta_date: datetime.date):
        """
        محاسبه سود تراکنش
        """
        start_date: datetime.date = az_date
        end_date: datetime.date = ta_date
        if self.effective_date > az_date:
            start_date = self.effective_date  # این واریزی در بین دوره محاسبه سود انجام شده نه از پیش از آن بنابراین سود معادل تعداد روز شامل را در یافت میکند
        # تا زمان درست شدن تاریخ انقضای تراکنش ها این شرط فیر فعال میگردد
        # if self.expire_date < ta_date:
        #     end_date = self.expire_date  # تاریخ پایان قرارداد این واریزی پیش از پایان بازه محاسبه سود است بنابراین سود تعداد روز شامل را دریافت میکند

        # sod = tr.percent  # نحوه محاسبه سود؟؟؟؟؟؟؟؟ اینجاست
        sod = self.percent_calculator()  # نحوه محاسبه سود؟؟؟؟؟؟؟؟ اینجاست
        mohasebe_sod: ProfitCalculate = ProfitCalculate()
        mohasebe_sod.transaction = self
        mohasebe_sod.Profile = self.profile
        mohasebe_sod.kind_id = 1
        mohasebe_sod.date_from = start_date
        mohasebe_sod.date_to = end_date
        mohasebe_sod.days = (end_date - start_date).days + 1  # فاصله روز شروع تا پایان +۱ شد
        mohasebe_sod.amount = self.amount
        mohasebe_sod.percent = sod
        mohasebe_sod.calculated_amount = round(self.amount * (sod / 10000000) * (mohasebe_sod.days / day_in_month), 0)
        if self.kind_id == 1:  # این تراکنش از نوع سپرده گزاری است
            pass
        if self.kind_id == 2:  # این تراکنش از نوع مرجوعی است
            mohasebe_sod.calculated_amount = -1 * mohasebe_sod.calculated_amount
        return mohasebe_sod

    def presenter_profit_calculator(self, az_date: datetime.date, ta_date: datetime.date):
        """
        محاسبه سود معرف
        برای تمام سپرده گزاری های معرف شده در ۹۳ روز پس از اولین واریزی به مدت ۶ ماه سود معرفی به معرف تعلق میگیرد
        """
        presenter = Profile.objects.get(id=self.profile.presenter.id)
        start_date: datetime.date = az_date
        end_date: datetime.date = ta_date
        if self.effective_date > az_date:
            start_date = self.effective_date  # این واریزی در بین دوره محاسبه سود انجام شده نه از پیش از آن بنابراین سود معادل تعداد روز شامل را در یافت میکند
        # تا زمان درست شدن تاریخ انقضای تراکنش ها این شرط فیر فعال شده و شرط زیر جایگزین میگردد
        # if self.expire_date < ta_date:
        #     end_date = self.expire_date  # تاریخ پایان قرارداد این واریزی پیش از پایان بازه محاسبه سود است بنابراین سود تعداد روز شامل را دریافت میکند
        #     return 0

        # فقط به واریزی های ۹۳ روز اول سود معرف تعلق میگرد
        # if self.profile.first_transaction.effective_date + datetime.timedelta(days=93) > self.effective_date:
        #     #     این واریزی در بازه ۹۳ روز اولیه سپرده گزار نیست بنابراین سودی به معرف تعلق نمیگیرد
        #     return 0

        if self.effective_date + datetime.timedelta(day_for_calculate_presenter_profit) < ta_date:
            end_date = self.effective_date + datetime.timedelta(
                day_for_calculate_presenter_profit)  # تاریخ پایان قرارداد این واریزی پیش از پایان بازه محاسبه سود است بنابراین سود تعداد روز شامل را دریافت میکند
        if self.effective_date + datetime.timedelta(day_for_calculate_presenter_profit) < az_date:
            end_date = start_date  # اگر تاریخ موثر قدیمی تر از بازه تعریف شده برای دریافت سود معرف بود به معرفی سودی تعلق نمیگیرد

        # sod = tr.percent  # نحوه محاسبه سود؟؟؟؟؟؟؟؟ اینجاست
        sod = self.presenter_percent_calculator(ta_date=ta_date)  # نحوه محاسبه سود؟؟؟؟؟؟؟؟ اینجاست
        mohasebe_sod: ProfitCalculate = ProfitCalculate()
        mohasebe_sod.transaction = self
        mohasebe_sod.Profile = presenter
        mohasebe_sod.kind_id = 2
        mohasebe_sod.date_from = start_date
        mohasebe_sod.date_to = end_date
        mohasebe_sod.days = (end_date - start_date).days + 1  # فاصله روز شروع تا پایان +۱ شد
        mohasebe_sod.amount = self.amount
        mohasebe_sod.percent = sod
        mohasebe_sod.calculated_amount = round(self.amount * (sod / 10000000) * (mohasebe_sod.days / day_in_month), 0)
        mohasebe_sod.description = f'محاسبه سود معرفی {self.profile}'
        return mohasebe_sod


class ProfitKind(models.Model):
    name = models.CharField(max_length=20, null=True, blank=True)

    def __str__(self):
        return self.name


class ProfitCalculate(models.Model):
    """
    محاسبه سود
    """
    # user = models.ForeignKey(get_user_model(), verbose_name='کاربر', on_delete=models.DO_NOTHING,
    #                          related_name='ProfitCalculates', blank=False, null=False)
    Profile = models.ForeignKey(Profile, on_delete=models.DO_NOTHING, related_name='profiteCalculates', null=False,
                                blank=False)
    date_from = models.DateField(verbose_name='از تاریخ', blank=False, null=False)
    date_to = models.DateField(verbose_name='تا تاریخ', blank=True, null=True)
    days = models.PositiveIntegerField(verbose_name='تعداد روز', blank=True, null=True)
    amount = models.PositiveBigIntegerField(verbose_name='مبلغ', blank=True, null=True)
    percent = models.IntegerField(verbose_name='درصد', validators=[MinValueValidator(1), MaxValueValidator(10)],
                                  blank=True, null=True)
    # تراکنشی که این رکورد بر اساس آن ساخته شده
    transaction = models.OneToOneField(to=Transaction, verbose_name='تراکنش متناظر', blank=False, null=False,
                                       on_delete=models.DO_NOTHING,
                                       related_name='ProfitCalculates')
    calculated_amount = models.PositiveIntegerField(verbose_name='مبلغ سود محاسبه شده', null=True, blank=True)
    balance = models.BooleanField(verbose_name='تسویه شده است', default=False)
    # تراکنشی که باغث شده این رکورد سود محاسبه شده تسویه گردد
    transaction_balance = models.ForeignKey(to=Transaction, verbose_name='تسویه شده با تراکنش',
                                            related_name='ProfitCalculate', on_delete=models.DO_NOTHING, null=True,
                                            blank=True)
    kind = models.ForeignKey(ProfitKind, on_delete=models.DO_NOTHING, null=False, blank=False)
    description = models.TextField(null=True, blank=True)
    comment = models.TextField(null=True, blank=True)
    created = models.DateTimeField(auto_now_add=True, editable=False)

    class Meta:
        ordering = ['created']

    def __str__(self):
        return f'for:({self.transaction}) from:({self.date_from} = {m2sh(self.date_from)}) to:({self.date_to} = {m2sh(self.date_to)}) days: ({self.days}) ' \
               f'amount: ({self.amount})  percent: ({self.percent}) final: ({self.calculated_amount})'


class PelekanKind(models.Model):
    title = models.CharField(max_length=200)
    start_date = models.DateField(null=True, blank=True)
    description = models.TextField(null=True, blank=True)

    def __str__(self):
        return self.title


class Pelekan(models.Model):
    az = models.PositiveBigIntegerField()
    ta = models.PositiveBigIntegerField()
    percent = models.IntegerField()
    kind = models.ForeignKey(PelekanKind, null=True, blank=True, on_delete=models.DO_NOTHING, related_name='kinds')
    description = models.TextField(null=True, blank=True)

    def __str__(self):
        return f'{self.kind.title}: {self.az} --- {self.ta}  :  {self.percent}'


class Post(models.Model):
    title = models.CharField('عنوان', max_length=100, blank=False, null=False)
    body = models.TextField('متن', blank=True, null=True)
    is_public = models.BooleanField('عمومی؟', default=True, )
    owner = models.ForeignKey(get_user_model(), on_delete=models.CASCADE, related_name='posts')

    def __str__(self):
        return f'{self.title}'


def mohasebe_sod_all(az_date: datetime, ta_date: datetime):
    profit_excel = openpyxl.Workbook()
    profit_sheet = profit_excel.active
    profit_sheet.title = "sadid Sheet1"

    profit_sheet.cell(row=1, column=1, value="row num")
    profit_sheet.cell(row=1, column=2, value="id")
    profit_sheet.cell(row=1, column=3, value="name")
    profit_sheet.cell(row=1, column=4, value="amount")
    profit_sheet.cell(row=1, column=5, value="hesab")
    profit_sheet.cell(row=1, column=6, value="محاسبه شده توسط نرم افزار قدیمی")
    profit_sheet.cell(row=1, column=7, value="اختلاف")

    profit_details_excel = openpyxl.Workbook()
    profit_details_sheet = profit_details_excel.active
    profit_details_sheet.title = "sadid Sheet1"

    profit_details_sheet.cell(row=1, column=1, value="row num")
    profit_details_sheet.cell(row=1, column=2, value="Profile_id")
    profit_details_sheet.cell(row=1, column=3, value="Profile_name")
    profit_details_sheet.cell(row=1, column=4, value="transaction_id")
    profit_details_sheet.cell(row=1, column=5, value="percent")
    profit_details_sheet.cell(row=1, column=6, value="amount")
    profit_details_sheet.cell(row=1, column=7, value="transaction_date")
    profit_details_sheet.cell(row=1, column=8, value="date_from")
    profit_details_sheet.cell(row=1, column=9, value="date_to")
    profit_details_sheet.cell(row=1, column=10, value="days")
    profit_details_sheet.cell(row=1, column=11, value="calculated_amount")
    profit_details_sheet.cell(row=1, column=12, value="تاریخ")

    counter = detail_counter = 1
    for pr in Profile.objects.all():
        p: Profile = pr
        try:
            sod_list, sod_sum = mohasebe_sod_1_nafar(p.id, az_date, ta_date)

            counter += 1
            profit_sheet.cell(row=counter, column=1, value=counter - 1)
            profit_sheet.cell(row=counter, column=2, value=f"{p.id}")
            profit_sheet.cell(row=counter, column=3, value=f"{p.first_name} {p.last_name}")
            profit_sheet.cell(row=counter, column=4, value=sod_sum)
            profit_sheet.cell(row=counter, column=5, value=p.shomare_hesab)

            old_calculated_value = 0
            try:
                old_calculated: Transaction = Transaction.objects.filter(profile=pr, kind_id=3, effective_date=ta_date)[
                    0]
                old_calculated_value = old_calculated.amount
            except IndexError:
                pass

            profit_sheet.cell(row=counter, column=6, value=old_calculated_value)
            profit_sheet.cell(row=counter, column=7, value=sod_sum - old_calculated_value)

            for pc in sod_list:
                detail_counter = detail_counter + 1
                pc: ProfitCalculate = pc
                profit_details_sheet.cell(row=detail_counter, column=1, value=detail_counter - 1)
                profit_details_sheet.cell(row=detail_counter, column=2, value=f"{pc.Profile.id}")
                profit_details_sheet.cell(row=detail_counter, column=3,
                                          value=f"{pc.Profile.first_name} {pc.Profile.last_name}")
                profit_details_sheet.cell(row=detail_counter, column=4, value=f"{pc.transaction.id}")
                profit_details_sheet.cell(row=detail_counter, column=5, value=f"{pc.percent}")
                profit_details_sheet.cell(row=detail_counter, column=6, value=pc.amount)
                profit_details_sheet.cell(row=detail_counter, column=7, value=f"{pc.transaction.effective_date}")
                profit_details_sheet.cell(row=detail_counter, column=8, value=f"{pc.date_from}")
                profit_details_sheet.cell(row=detail_counter, column=9, value=f"{pc.date_to}")
                profit_details_sheet.cell(row=detail_counter, column=10, value=f"{pc.days}")
                profit_details_sheet.cell(row=detail_counter, column=11, value=pc.calculated_amount)
                profit_details_sheet.cell(row=detail_counter, column=12, value=f"{m2sh(pc.transaction.effective_date)}")

        except Pelekan.MultipleObjectsReturned:
            print(f'ERROR for {p}({p.id})  MultipleObjectsReturned')
            sod_sum = f'ERROR for {p}({p.id})  MultipleObjectsReturned'
        except Pelekan.DoesNotExist:
            print(f'ERROR for {p}({p.id})  DoesNotExist')
            sod_sum = f'ERROR for {p}({p.id})  DoesNotExist'

    profit_excel.save(filename=f"profit-{jdatetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")
    profit_details_excel.save(filename=f"profit_details-{jdatetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")
    print(f'finish')


def mohasebe_sod_moarefi_all(az_date: datetime, ta_date: datetime):
    profit_excel = openpyxl.Workbook()
    profit_sheet = profit_excel.active
    profit_sheet.title = "sadid Sheet1"

    profit_sheet.cell(row=1, column=1, value="row num")
    profit_sheet.cell(row=1, column=2, value="id")
    profit_sheet.cell(row=1, column=3, value="name")
    profit_sheet.cell(row=1, column=4, value="amount")
    profit_sheet.cell(row=1, column=5, value="hesab")
    profit_sheet.cell(row=1, column=6, value="محاسبه شده توسط نرم افزار قدیمی")
    profit_sheet.cell(row=1, column=7, value="اختلاف")

    profit_details_excel = openpyxl.Workbook()
    profit_details_sheet = profit_details_excel.active
    profit_details_sheet.title = "sadid Sheet1"

    profit_details_sheet.cell(row=1, column=1, value="row num")
    profit_details_sheet.cell(row=1, column=2, value="Profile_id")
    profit_details_sheet.cell(row=1, column=3, value="Profile_name")
    profit_details_sheet.cell(row=1, column=4, value="transaction_id")
    profit_details_sheet.cell(row=1, column=5, value="percent")
    profit_details_sheet.cell(row=1, column=6, value="amount")
    profit_details_sheet.cell(row=1, column=7, value="transaction_date")
    profit_details_sheet.cell(row=1, column=8, value="date_from")
    profit_details_sheet.cell(row=1, column=9, value="date_to")
    profit_details_sheet.cell(row=1, column=10, value="days")
    profit_details_sheet.cell(row=1, column=11, value="calculated_amount")
    profit_details_sheet.cell(row=1, column=12, value="تاریخ")
    profit_details_sheet.cell(row=1, column=13, value="name")

    counter = detail_counter = 1
    for pr in Profile.objects.all():
        p: Profile = pr
        try:
            sod_list, sod_sum = mohasebe_sod_moarefi_1_nafar(p.id, az_date, ta_date)

            counter += 1
            profit_sheet.cell(row=counter, column=1, value=counter - 1)
            profit_sheet.cell(row=counter, column=2, value=f"{p.id}")
            profit_sheet.cell(row=counter, column=3, value=f"{p.first_name} {p.last_name}")
            profit_sheet.cell(row=counter, column=4, value=f"{sod_sum}")
            profit_sheet.cell(row=counter, column=5, value=p.shomare_hesab)
            old_calculated_value = 0
            try:
                old_calculated: Transaction = Transaction.objects.filter(profile=pr, kind_id=5,
                                                                         effective_date=ta_date).aggregate(
                    Sum('amount'))
                old_calculated_value = old_calculated['amount__sum'] or 0
            except IndexError:
                pass

            profit_sheet.cell(row=counter, column=6, value=old_calculated_value)
            profit_sheet.cell(row=counter, column=7, value=sod_sum - old_calculated_value)

            for pc in sod_list:
                detail_counter = detail_counter + 1
                pc: ProfitCalculate = pc
                profit_details_sheet.cell(row=detail_counter, column=1, value=detail_counter - 1)
                profit_details_sheet.cell(row=detail_counter, column=2, value=f"{pc.Profile.id}")
                profit_details_sheet.cell(row=detail_counter, column=3,
                                          value=f"{pc.Profile.first_name} {pc.Profile.last_name}")
                profit_details_sheet.cell(row=detail_counter, column=4, value=f"{pc.transaction.id}")
                profit_details_sheet.cell(row=detail_counter, column=5, value=f"{pc.percent}")
                profit_details_sheet.cell(row=detail_counter, column=6, value=f"{pc.amount}")
                profit_details_sheet.cell(row=detail_counter, column=7, value=f"{pc.transaction.effective_date}")
                profit_details_sheet.cell(row=detail_counter, column=8, value=f"{pc.date_from}")
                profit_details_sheet.cell(row=detail_counter, column=9, value=f"{pc.date_to}")
                profit_details_sheet.cell(row=detail_counter, column=10, value=f"{pc.days}")
                profit_details_sheet.cell(row=detail_counter, column=11, value=f"{pc.calculated_amount}")
                profit_details_sheet.cell(row=detail_counter, column=12, value=f"{m2sh(pc.transaction.effective_date)}")
                profit_details_sheet.cell(row=detail_counter, column=13,
                                          value=f"{pc.transaction.profile.first_name} {pc.transaction.profile.last_name}")

        except Pelekan.MultipleObjectsReturned:
            print(f'ERROR for {p}({p.id})  MultipleObjectsReturned')
            sod_sum = f'ERROR for {p}({p.id})  MultipleObjectsReturned'
        except Pelekan.DoesNotExist:
            print(f'ERROR for {p}({p.id})  DoesNotExist')
            sod_sum = f'ERROR for {p}({p.id})  DoesNotExist'

    profit_excel.save(filename=f"profit-moarefi-{jdatetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")
    profit_details_excel.save(
        filename=f"profit-moarefi_details-{jdatetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")
    print(f'finish')


def mohasebe_sod_1_nafar(a: int, az_date: datetime, ta_date: datetime):
    m = Profile.objects.get(id=a)
    sod_list, sod_sum = m.mohasebe_sod_old(az_date=az_date, ta_date=ta_date)
    return sod_list, sod_sum


def mohasebe_sod_moarefi_1_nafar(a: int, az_date: datetime, ta_date: datetime):
    m = Profile.objects.get(id=a)
    sod_list, sod_sum = m.mohasebe_sod_moarefi(az_date=az_date, ta_date=ta_date)
    return sod_list, sod_sum


def shamsi_to_miladi(j_date: str, sep: str = "/") -> datetime.date:
    """
    تبدیل تاریخ شمسی به میلادی
    :param j_date: رشته تاریخ شمسی مانند ۱۴۰۰/۱۰/۰۳
    :param sep: جدا کننده تاریخ. پیشفرض / است
    :return: datetime.date تاریخ میلادی
    """
    shamsi = j_date.split(sep=sep)
    jyear = int(float(shamsi[0]))
    jmonth = int(float(shamsi[1]))
    jday = int(float(shamsi[2]))
    # jyear = int(float(j_date[0:4]))
    # jmonth = int(float(j_date[5:7]))
    # jday = int(float(j_date[8:10]))
    (gyear, gmonth, gday) = jdatetime.JalaliToGregorian(jyear=jyear, jmonth=jmonth,
                                                        jday=jday).getGregorianList()
    return datetime.date(gyear, gmonth, gday)


def sh2m(j_date: str, sep: str = "/") -> datetime.date:
    """
    تبدیل تاریخ شمسی به میلادی
    :param j_date: رشته تاریخ شمسی مانند ۱۴۰۰/۱۰/۰۳
    :param sep: جدا کننده تاریخ. پیشفرض / است
    :return: datetime.date تاریخ میلادی
    """
    return shamsi_to_miladi(j_date, sep)


def miladi_to_shamsi(g_date: datetime.date) -> str:
    return jdatetime.date.fromgregorian(date=g_date).strftime(format="%Y-%m-%d")


def m2sh(g_date: datetime.date) -> str:
    return miladi_to_shamsi(g_date)
