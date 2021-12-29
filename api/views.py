import datetime

import jdatetime
import openpyxl
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum
from django.shortcuts import render
from graphene_django.views import GraphQLView

from api.models import Profile, mohasebe_sod_1_nafar, Transaction, ProfitCalculate, Pelekan, m2sh, \
    mohasebe_sod_moarefi_1_nafar
from backend import settings


class PrivateGraphQLView(LoginRequiredMixin, GraphQLView):
    pass


def home(request):
    return render(request, 'home.html')


def mohasebe_sod_all_export_excel(az_date: datetime, ta_date: datetime):
    profit_excel = openpyxl.Workbook()
    profit_sheet = profit_excel.active
    profit_sheet.title = "sadid Sheet1"

    profit_sheet.cell(row=1, column=1, value="row num")
    profit_sheet.cell(row=1, column=2, value="id")
    profit_sheet.cell(row=1, column=3, value="name")
    profit_sheet.cell(row=1, column=4, value="amount")
    profit_sheet.cell(row=1, column=5, value="hesab")
    profit_sheet.cell(row=1, column=6, value="محاسبه شده توسط نرم افزار قدیمی")
    profit_sheet.cell(row=1, column=7, value="اختلاف")

    counter = detail_counter = 1
    for pr in Profile.objects.all():
        p: Profile = pr
        try:
            sod_list, sod_sum = mohasebe_sod_1_nafar(p.id, az_date, ta_date)

            counter += 1
            profit_sheet.cell(row=counter, column=1, value=counter - 1)
            profit_sheet.cell(row=counter, column=2, value=f"{p.id}")
            profit_sheet.cell(row=counter, column=3, value=f"{p.first_name} {p.last_name}")
            profit_sheet.cell(row=counter, column=4, value=sod_sum)
            profit_sheet.cell(row=counter, column=5, value=p.account_number)

            old_calculated_value = 0
            try:
                old_calculated: Transaction = Transaction.objects.filter(profile=pr, kind_id=3, effective_date=ta_date)[
                    0]
                old_calculated_value = old_calculated.amount
            except IndexError:
                pass

            profit_sheet.cell(row=counter, column=6, value=old_calculated_value)
            profit_sheet.cell(row=counter, column=7, value=sod_sum - old_calculated_value)



        except Pelekan.MultipleObjectsReturned:
            print(f'ERROR for {p}({p.id})  MultipleObjectsReturned')
            sod_sum = f'ERROR for {p}({p.id})  MultipleObjectsReturned'
        except Pelekan.DoesNotExist:
            print(f'ERROR for {p}({p.id})  DoesNotExist')
            sod_sum = f'ERROR for {p}({p.id})  DoesNotExist'
    filename = f"profit_calc/profit-{jdatetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    profit_excel.save(filename=f"{settings.MEDIA_ROOT}/{filename}")
    print(f'finish')
    return f"{settings.MEDIA_URL}{filename}"


def mohasebe_sod_detail_all_export_excel(az_date: datetime, ta_date: datetime):
    profit_excel = openpyxl.Workbook()
    profit_sheet = profit_excel.active
    profit_sheet.title = "sadid Sheet1"

    profit_sheet.cell(row=1, column=1, value="row num")
    profit_sheet.cell(row=1, column=2, value="id")
    profit_sheet.cell(row=1, column=3, value="name")
    profit_sheet.cell(row=1, column=4, value="amount")
    profit_sheet.cell(row=1, column=5, value="hesab")
    profit_sheet.cell(row=1, column=6, value="محاسبه شده توسط نرم افزار قدیمی")
    profit_sheet.cell(row=1, column=7, value="اختلاف")

    profit_details_excel = openpyxl.Workbook()
    profit_details_sheet = profit_details_excel.active
    profit_details_sheet.title = "sadid Sheet1"

    profit_details_sheet.cell(row=1, column=1, value="row num")
    profit_details_sheet.cell(row=1, column=2, value="Profile_id")
    profit_details_sheet.cell(row=1, column=3, value="Profile_name")
    profit_details_sheet.cell(row=1, column=4, value="transaction_id")
    profit_details_sheet.cell(row=1, column=5, value="percent")
    profit_details_sheet.cell(row=1, column=6, value="amount")
    profit_details_sheet.cell(row=1, column=7, value="transaction_date")
    profit_details_sheet.cell(row=1, column=8, value="date_from")
    profit_details_sheet.cell(row=1, column=9, value="date_to")
    profit_details_sheet.cell(row=1, column=10, value="days")
    profit_details_sheet.cell(row=1, column=11, value="calculated_amount")
    profit_details_sheet.cell(row=1, column=12, value="تاریخ")

    counter = detail_counter = 1
    for pr in Profile.objects.all():
        p: Profile = pr
        try:
            sod_list, sod_sum = mohasebe_sod_1_nafar(p.id, az_date, ta_date)

            counter += 1
            profit_sheet.cell(row=counter, column=1, value=counter - 1)
            profit_sheet.cell(row=counter, column=2, value=f"{p.id}")
            profit_sheet.cell(row=counter, column=3, value=f"{p.first_name} {p.last_name}")
            profit_sheet.cell(row=counter, column=4, value=sod_sum)
            profit_sheet.cell(row=counter, column=5, value=p.account_number)

            old_calculated_value = 0
            try:
                old_calculated: Transaction = Transaction.objects.filter(profile=pr, kind_id=3, effective_date=ta_date)[
                    0]
                old_calculated_value = old_calculated.amount
            except IndexError:
                pass

            profit_sheet.cell(row=counter, column=6, value=old_calculated_value)
            profit_sheet.cell(row=counter, column=7, value=sod_sum - old_calculated_value)

            for pc in sod_list:
                detail_counter = detail_counter + 1
                pc: ProfitCalculate = pc
                profit_details_sheet.cell(row=detail_counter, column=1, value=detail_counter - 1)
                profit_details_sheet.cell(row=detail_counter, column=2, value=f"{pc.Profile.id}")
                profit_details_sheet.cell(row=detail_counter, column=3,
                                          value=f"{pc.Profile.first_name} {pc.Profile.last_name}")
                profit_details_sheet.cell(row=detail_counter, column=4, value=f"{pc.transaction.id}")
                profit_details_sheet.cell(row=detail_counter, column=5, value=f"{pc.percent}")
                profit_details_sheet.cell(row=detail_counter, column=6, value=pc.amount)
                profit_details_sheet.cell(row=detail_counter, column=7, value=f"{pc.transaction.effective_date}")
                profit_details_sheet.cell(row=detail_counter, column=8, value=f"{pc.date_from}")
                profit_details_sheet.cell(row=detail_counter, column=9, value=f"{pc.date_to}")
                profit_details_sheet.cell(row=detail_counter, column=10, value=f"{pc.days}")
                profit_details_sheet.cell(row=detail_counter, column=11, value=pc.calculated_amount)
                profit_details_sheet.cell(row=detail_counter, column=12, value=f"{m2sh(pc.transaction.effective_date)}")

        except Pelekan.MultipleObjectsReturned:
            print(f'ERROR for {p}({p.id})  MultipleObjectsReturned')
            sod_sum = f'ERROR for {p}({p.id})  MultipleObjectsReturned'
        except Pelekan.DoesNotExist:
            print(f'ERROR for {p}({p.id})  DoesNotExist')
            sod_sum = f'ERROR for {p}({p.id})  DoesNotExist'

    profit_excel.save(filename=f"profit-{jdatetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")
    profit_details_excel.save(filename=f"profit_details-{jdatetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")
    print(f'finish')


def mohasebe_sod_moarefi_all_export_excel(az_date: datetime, ta_date: datetime):
    profit_excel = openpyxl.Workbook()
    profit_sheet = profit_excel.active
    profit_sheet.title = "sadid Sheet1"

    profit_sheet.cell(row=1, column=1, value="row num")
    profit_sheet.cell(row=1, column=2, value="id")
    profit_sheet.cell(row=1, column=3, value="name")
    profit_sheet.cell(row=1, column=4, value="amount")
    profit_sheet.cell(row=1, column=5, value="hesab")
    profit_sheet.cell(row=1, column=6, value="محاسبه شده توسط نرم افزار قدیمی")
    profit_sheet.cell(row=1, column=7, value="اختلاف")

    profit_details_excel = openpyxl.Workbook()
    profit_details_sheet = profit_details_excel.active
    profit_details_sheet.title = "sadid Sheet1"

    profit_details_sheet.cell(row=1, column=1, value="row num")
    profit_details_sheet.cell(row=1, column=2, value="Profile_id")
    profit_details_sheet.cell(row=1, column=3, value="Profile_name")
    profit_details_sheet.cell(row=1, column=4, value="transaction_id")
    profit_details_sheet.cell(row=1, column=5, value="percent")
    profit_details_sheet.cell(row=1, column=6, value="amount")
    profit_details_sheet.cell(row=1, column=7, value="transaction_date")
    profit_details_sheet.cell(row=1, column=8, value="date_from")
    profit_details_sheet.cell(row=1, column=9, value="date_to")
    profit_details_sheet.cell(row=1, column=10, value="days")
    profit_details_sheet.cell(row=1, column=11, value="calculated_amount")
    profit_details_sheet.cell(row=1, column=12, value="تاریخ")
    profit_details_sheet.cell(row=1, column=13, value="name")

    counter = detail_counter = 1
    for pr in Profile.objects.all():
        p: Profile = pr
        try:
            sod_list, sod_sum = mohasebe_sod_moarefi_1_nafar(p.id, az_date, ta_date)

            counter += 1
            profit_sheet.cell(row=counter, column=1, value=counter - 1)
            profit_sheet.cell(row=counter, column=2, value=f"{p.id}")
            profit_sheet.cell(row=counter, column=3, value=f"{p.first_name} {p.last_name}")
            profit_sheet.cell(row=counter, column=4, value=f"{sod_sum}")
            profit_sheet.cell(row=counter, column=5, value=p.accountـnumber)
            old_calculated_value = 0
            try:
                old_calculated: Transaction = Transaction.objects.filter(profile=pr, kind_id=5,
                                                                         effective_date=ta_date).aggregate(
                    Sum('amount'))
                old_calculated_value = old_calculated['amount__sum'] or 0
            except IndexError:
                pass

            profit_sheet.cell(row=counter, column=6, value=old_calculated_value)
            profit_sheet.cell(row=counter, column=7, value=sod_sum - old_calculated_value)

            for pc in sod_list:
                detail_counter = detail_counter + 1
                pc: ProfitCalculate = pc
                profit_details_sheet.cell(row=detail_counter, column=1, value=detail_counter - 1)
                profit_details_sheet.cell(row=detail_counter, column=2, value=f"{pc.Profile.id}")
                profit_details_sheet.cell(row=detail_counter, column=3,
                                          value=f"{pc.Profile.first_name} {pc.Profile.last_name}")
                profit_details_sheet.cell(row=detail_counter, column=4, value=f"{pc.transaction.id}")
                profit_details_sheet.cell(row=detail_counter, column=5, value=f"{pc.percent}")
                profit_details_sheet.cell(row=detail_counter, column=6, value=f"{pc.amount}")
                profit_details_sheet.cell(row=detail_counter, column=7, value=f"{pc.transaction.effective_date}")
                profit_details_sheet.cell(row=detail_counter, column=8, value=f"{pc.date_from}")
                profit_details_sheet.cell(row=detail_counter, column=9, value=f"{pc.date_to}")
                profit_details_sheet.cell(row=detail_counter, column=10, value=f"{pc.days}")
                profit_details_sheet.cell(row=detail_counter, column=11, value=f"{pc.calculated_amount}")
                profit_details_sheet.cell(row=detail_counter, column=12, value=f"{m2sh(pc.transaction.effective_date)}")
                profit_details_sheet.cell(row=detail_counter, column=13,
                                          value=f"{pc.transaction.profile.first_name} {pc.transaction.profile.last_name}")

        except Pelekan.MultipleObjectsReturned:
            print(f'ERROR for {p}({p.id})  MultipleObjectsReturned')
            sod_sum = f'ERROR for {p}({p.id})  MultipleObjectsReturned'
        except Pelekan.DoesNotExist:
            print(f'ERROR for {p}({p.id})  DoesNotExist')
            sod_sum = f'ERROR for {p}({p.id})  DoesNotExist'

    profit_excel.save(filename=f"profit-moarefi-{jdatetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")
    profit_details_excel.save(
        filename=f"profit-moarefi_details-{jdatetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")
    print(f'finish')


def excel_to_bank_export_paya(in_file, out_file, sum, count):
    """
    تبدیل خروجی نرم افزار مرادعلیان به فرمت مورد نیاز بانک برای پایا
    :param in_file: فایل اکسل مرادعلیان
    :param out_file: فایل خروجی مورد نظر با پسوند ccti
    :param sum: مجموع پول وارد شده در فایل
    :param count: تعداد واریزی های وارد شده در فایل
    :return:
    """
    # path = "/Users/alamalhoda/Projects/AzarSaryame/backend/api/scripts/"
    inp = f'{in_file}'
    # outp = f'{path}paya-new.ccti'
    outp = out_file
    wb_obj = openpyxl.load_workbook(inp)
    sheet_obj = wb_obj.active

    max_row_no = sheet_obj.max_row  # تعداد ردیف ها
    max_col = sheet_obj.max_column  # تعداد ستون ها
    print(max_row_no)
    export_file = open(outp, "a")
    header = (f'<?xml version="1.0" encoding="utf-8" standalone="yes"?>\n'
              f'<Document xmlns="urn:iso:std:iso:20022:tech:xsd:pain.002.001.03">\n'
              f'  <CstmrCdtTrfInitn>\n'
              f'    <GrpHdr>\n'
              f'      <MsgId>IR100600521070009334410001000000000</MsgId>\n'
              f'      <CreDtTm>1400-010-02T14:44:00</CreDtTm>\n'
              f'      <NbOfTxs>{count}</NbOfTxs>\n'
              f'      <CtrlSum>{sum}</CtrlSum>\n'
              f'      <InitgPty>\n'
              f'        <Nm>سامان مزرعه لی</Nm>\n'
              f'      </InitgPty>\n'
              f'    </GrpHdr>\n'
              f'    <PmtInf>\n'
              f'      <PmtInfId>1</PmtInfId>\n'
              f'      <PmtMtd Ccy="IRR">TRF</PmtMtd>\n'
              f'      <NbOfTxs>{count}</NbOfTxs>\n'
              f'      <CtrlSum>{sum}</CtrlSum>\n'
              f'      <ReqdExctnDt>1400-10-08</ReqdExctnDt>\n'
              f'      <Dbtr>\n'
              f'        <Nm>سامان مزرعه لی</Nm>\n'
              f'      </Dbtr>\n'
              f'      <DbtrAcct>\n'
              f'        <Id>\n'
              f'          <IBAN>IR100600521070009334410001</IBAN>\n'
              f'        </Id>\n'
              f'      </DbtrAcct>\n'
              f'      <DbtrAgt>\n'
              f'        <FinInstnId>\n'
              f'          <BIC>BMJIIRTHXXX</BIC>\n'
              f'        </FinInstnId>\n'
              f'      </DbtrAgt>')

    export_file.write(header)

    for i in range(2, max_row_no + 1):
        first_name = sheet_obj.cell(row=i, column=3).value  # نام
        last_name = sheet_obj.cell(row=i, column=4).value  # famil
        IBAN = sheet_obj.cell(row=i, column=5).value  # IBAN شماره حساب
        InstdAmt = sheet_obj.cell(row=i, column=6).value  # InstdAmt مبلغ
        # print(last_name)

        xml = (
            f'<CdtTrfTxInf>\n'
            f'  <PmtId>\n'
            f'    <InstrId>EMPTY</InstrId>\n'
            f'    <EndToEndId>EMPTY</EndToEndId>\n'
            f'  </PmtId>\n'
            f'  <Amt>\n'
            f'    <InstdAmt Ccy="IRR">{InstdAmt}</InstdAmt>\n'
            f'  </Amt>\n'
            f'  <Cdtr>\n'
            f'    <Nm>{first_name} {last_name}</Nm>\n'
            f'    <Id>\n'
            f'      <PrvtId>\n'
            f'        <Othr>\n'
            f'        <Id>EMPTY</Id>\n'
            f'        </Othr>\n'
            f'      </PrvtId>\n'
            f'     </Id>\n'
            f'  </Cdtr>\n'
            f'  <CdtrAcct>\n'
            f'    <Id>\n'
            f'      <IBAN>{IBAN}</IBAN>\n'
            f'    </Id>\n'
            f'  </CdtrAcct>\n'
            f'</CdtTrfTxInf>\n')
        export_file.write(xml)

    footer = (f'    </PmtInf>\n'
              f'  </CstmrCdtTrfInitn>\n'
              f'</Document>')

    export_file.write(footer)
